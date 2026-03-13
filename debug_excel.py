import openpyxl
import sys

file_path = 'VENDOR DATA PART.xlsx'

try:
    print(f"Opening {file_path} in read-only mode...")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"Reported Max Row: {max_row}")
    print(f"Reported Max Col: {max_col}")
    
    # Iterate to count actual rows if max_row is unreliable or huge
    count = 0
    for row in sheet.iter_rows():
        count += 1
        if count % 1000 == 0:
            print(f"Scanned {count} rows...")
        if count > 10000:
            print("Stopped scanning at 10,000 rows.")
            break
            
    print(f"Actual iterated rows: {count}")
    
    # Print header
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            headers.append(cell.value)
    print(f"Headers: {headers}")

except Exception as e:
    print(f"Error: {e}")
